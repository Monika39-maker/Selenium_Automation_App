from openpyxl import load_workbook, Workbook
from selenium import webdriver
import time

def read_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def write_excel(data, output_file):
    workbook = Workbook()
    sheet = workbook.active

    for row in data:
        sheet.append(row)

    workbook.save(output_file)

def check_facebook_friends(data):
    # Initialize Selenium WebDriver
    driver = webdriver.Chrome()  # You may need to change the WebDriver based on your browser
    driver.get("https://www.facebook.com/friends/list")  # Change this to the Facebook URL

    # Login to Facebook (You should replace 'email' and 'password' with your actual credentials)
    email_input = driver.find_element_by_id("email")
    email_input.send_keys("chayadangol@yahoo.com")

    password_input = driver.find_element_by_id("pass")
    password_input.send_keys("Monika14")
    password_input.submit()

    # Wait for the user to login manually (You may need to handle login automatically with more advanced techniques)

    # Iterate through the data and check if each name is in the friend list
    for row in data:
        name = row[0]
        
        # Example: Search for the name on Facebook (Assuming the search input field has ID "search_input_id")
        search_input = driver.find_element_by_id("search_input_id")
        search_input.clear()
        search_input.send_keys(name)
        search_input.submit()

        # Wait for the search results to load
        time.sleep(2)

        # Check if the user is found in the search results (Assuming the friend's profile has a class "friend-profile")
        friend_profile = driver.find_elements_by_class_name("friend-profile")
        if friend_profile:
            row[1] = "Yes"  # If found, mark as "Yes"
        else:
            row[1] = "No"   # If not found, mark as "No"

    # Close the browser window when finished
    driver.quit()

def main():
    # Read data from Excel
    excel_file = "Facebookfriendstest.xlsx"  # Change this to your Excel file path
    excel_data = read_excel(excel_file)

    # Check Facebook friends
    check_facebook_friends(excel_data)

    # Write updated data to a new Excel file
    output_file = "updated_data.xlsx"
    write_excel(excel_data, output_file)
    print(f"Updated data has been saved to {output_file}")

if __name__ == "__main__":
    main()
